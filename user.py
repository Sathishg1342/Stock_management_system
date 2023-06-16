
import openpyxl
wb=openpyxl.load_workbook('D:\\Sathish.G\\For Html\\for pyhton.xlsx')
#print(type(wb))
#sheets=wb.sheetnames
#print(sheets)
sh1 = wb['sample']
sh2 = wb['User']
#print(type(sh1))
#t=sh1.cell(3,2).value
#print(sh1.max_row,sh1.max_column)
'''
sh1.cell(5,1,value='physics')
sh1.cell(5,2,value=70)
sh1.cell(5,3,value=100)
wb.save('for python.xlsx')
'''
'''
n=int(input('enter the no of columns'))
for i in range(1,int(input('enter the no of rows:'))+1):
    for j in range(1,n+1):
        print('enter the value for cell(',i,',',j,'):')
        sh2.cell(i,j,value=input())

#sh1.delete_rows(1,sh1.max_row)

for i in range(1,sh2.max_row+1):
    print()
    for j in range(1,(sh2.max_column)+1):
        print(sh2.cell(i,j).value,end=' ')
'''

def login_check(username,password):
    flag=-1
    for i in range(1,sh1.max_row+1):
        if((sh1.cell(i,1).value==username) and (sh1.cell(i,2).value==password)):
            flag=1
            break
    else:
        for j in range(1,sh2.max_row+1):
            if(sh2.cell(j,1).value==username and sh2.cell(j,2).value==password):
                flag=0
                break
    return (flag)
#print(login_check('Admin2','psw2'))

def ShowUser():
    return ([[sh2.cell(i,j).value for j in range(1,sh2.max_column+1)]for i in range(1,sh2.max_row+1)])
#print(ShowUser())
'''
def adduser(username,password,name,phoneno):
    return add_user(username,password,name,phoneno)
'''
def add_user(username,password,name,phoneno):
    for i in range(1,sh2.max_row+1):
        if sh2.cell(i,1).value==username:
            return -1
        if(sh2.cell(i,1).value==None):
            mx=i-1
        else:
            mx=sh2.max_row
    sh2.cell(mx+1,1,value=username)
    sh2.cell(mx+1,2,value=password)
    sh2.cell(mx+1,3,value=name)
    sh2.cell(mx+1,4,value=phoneno)
    wb.save('D:\\Sathish.G\\For Html\\for pyhton.xlsx')
    if(mx+1==sh2.max_row):
        return 1
    else:
        return 0
#print(add_user('nithish@gmail.com','nsh','Nithish',98476846438))

def change_password(username,password):
    for i in range(1,sh2.max_row+1):
        if(sh2.cell(i,1).value==username):
            sh2.cell(i,2).value=password
            wb.save('D:\\Sathish.G\\For Html\\for pyhton.xlsx')
            return(1)
    else:
        return 0
#print(change_password('gowtham','gow'))
def remove_user(username):
    s=str(username)
    for i in range(1,sh2.max_row+1):
        if(sh2.cell(i,1).value==s):
            sh2.delete_rows(i)
            #print(0)
    wb.save('D:\\Sathish.G\\For Html\\for pyhton.xlsx')
#remove_user('nithish@gmail.com')
wb.save('D:\\Sathish.G\\For Html\\for pyhton.xlsx')
'''
n,m=3,4
print([[i*j for i in range(1,m+1)]for j in range(1,n+1)])
'''